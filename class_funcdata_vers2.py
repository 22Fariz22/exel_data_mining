# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import copy

class Exl_data:
    all_clients = 0

    def __init__(self,amount, name_file):  # обязательные  параметры при создании экземпляра
        self.amount = amount  # параметр для суммы клиентов
        self.name_file = name_file  # параметр для названия файла
        self.maxcol = 0  # макс.количство колонок в 1-ом листе
        self.maxrow = 0  # макс.количство строк в 1-ом листе
        self.count_row = 2  #
        self.find = 0  # сумма найденных строк
        self.clients = 0  # считаем сумму клиентов
        self.level = 0  # определим уровень комбинации поиска

        self.wb = load_workbook(filename=self.name_file)  # вставляем файл-эксель и присваеваем к wb
        self.sheet = self.wb.active.title  # текущий активный лист

        self.sh1 = self.wb['Сегменты']
        self.data2 = self.wb['Результат']['A1'].value  # 1-ый способ получить ячейку
        self.sh1.cell(6, 6).value  # 2-ой способ получить ячейку


        self.maxcol = self.sh1.max_column  # макс.количство колонок в 1-ом листе
        self.maxrow = self.sh1.max_row  # макс.количство строк в 1-ом листе
        print('amount- ',self.amount,', filename- ', self.name_file)
        print('maxcol- ',self.maxcol,', maxrow- ', self.maxrow)
        self.data_rows = []  # все строки 6-го уровня
        self.data_same_rows_6_lev = []  # все одинаковые строки 6-го уровня


    def __create_temp_sheet(self):
        self.wb.create_sheet(f'N={self.amount}')
        self.sh3 = self.wb[f'N={self.amount}']


    # итерирует все строки и отбирает по N-клиентов.создает отдельный лист
    def __find_amount_clients_in_all(self):

        for i in range(2, self.maxrow +1):  # итерируем по количество строкам листа Сегменты со 2 строки,тк 1 это названия колонок.
            if self.sh1.cell(row=i, column=self.maxcol).value >= self.amount:  # тут будет N-аргумент из функции
                self.find += 1  # считаем найденные подходящие строки
                self.clients += self.sh1.cell(row=i, column=self.maxcol).value  # считаем всех клиентов у подход. строк
                for j in range(1, self.maxcol + 1):  # итерируем по количеству колонок в строке таблицы Сегменты
                    c = self.sh1.cell(row=i, column=j)  # подходящую строку присваиваем переменной с
                    self.sh3.cell(row=self.count_row, column=j).value = c.value  # копируем строку в другую таблицу
                self.count_row += 1  # считаем

    # создаем список исходных строк
    def __create_clean_main_list(self):
        for row in self.sh1.iter_rows(min_row=2, max_row=self.maxrow,values_only=True):  # max_col это
            self.data_rows.append(list(row))


    # сохраняет  уник.строки и подсчитаны суммы клиентов.
    def __create_list_with_choose_col_and_uniq_rows_and_rewrite_to_new_sheet(self,lev,x):
        # data_same_rows = []
        list_col = []
        data_same_rows = copy.deepcopy(self.data_rows)  # копируем исходные в временный список

        if lev == 6:
            for i in data_same_rows: # закрываем х-колонку
                i[x] = 'XXX'
        if lev == 5:
            for i in data_same_rows: # закрываем х-колонку
                i[0] = 'XXX'
                i[x] = 'XXX'
        if lev == 4:
            for i in data_same_rows: # закрываем х-колонку
                i[0] = 'XXX'
                i[1] = 'XXX'
                i[x] = 'XXX'
        if lev == 3:
            for i in data_same_rows: # закрываем х-колонку
                i[0] = 'XXX'
                i[1] = 'XXX'
                i[2] = 'XXX'
                i[x] = 'XXX'
        if lev == 2:
            for i in data_same_rows: # закрываем х-колонку
                i[0] = 'XXX'
                i[1] = 'XXX'
                i[2] = 'XXX'
                i[3] = 'XXX'
                i[x] = 'XXX'


        for i in data_same_rows: # собирает из data_same_rows уникальные в list_col_0
            if i[0:6] not in list_col:
                list_col.append(i[0:6])


        # считаем клиентов всех одинаковых строк и вешаем на уникальную.
        for i in list_col:
            count_clients = 0
            for k in data_same_rows:
                if i == k[0:6]:
                    count_clients += k[6]
            i.append(count_clients)

        # отбиралка по кол-ву клиентов(amount)
        len_row = len(list_col)
        revers_row = list(reversed(range(len_row, 0, -1)))

        for spisok in reversed(range(0, len_row)):  # удаляю с конца чтобы удалять без изменений индекса строк
            if list_col[spisok][6] < self.amount:
                list_col.pop(spisok)


    # переписываем уник.строки из data_same_rows в новый лист
        count_row_sh3 = 2 # точка отсчета rows в новом листе
        self.wb.create_sheet(f'N={self.amount} lev={lev} col={x}')  # создаем новый лист
        temp_sh = self.wb[f'N={self.amount} lev={lev} col={x}']
        len_data = len(list_col)


        for spisok in range(0,len_data):
                for j in range(0, self.maxcol ):
                    cell = list_col[spisok][j]
                    temp_sh.cell(row=count_row_sh3, column=j+1).value = cell
                count_row_sh3 += 1

        #ставим 1ый ряд с именами колонок
        for i in range(1, self.maxcol + 1):
            pass

        print(self.wb[f'N={self.amount} lev={lev} col={x}'])


    def savefile(self):
        self.wb.save(f'amount_{self.amount}.xlsx')


    def level_6_all_cols(self):

        self.__create_temp_sheet() # лист для find_amount_clients_in_all
        self.__find_amount_clients_in_all()  # лист исходных данных с условием amount
        self.__create_clean_main_list()  # просто список всех исходных без условий. список data_rows
        for i in range(6):
            self.__create_list_with_choose_col_and_uniq_rows_and_rewrite_to_new_sheet(x=i,lev=6)

        for i in range(1,6):
            self.__create_list_with_choose_col_and_uniq_rows_and_rewrite_to_new_sheet(x=i,lev=5)

        for i in range(2,6):
            self.__create_list_with_choose_col_and_uniq_rows_and_rewrite_to_new_sheet(x=i,lev=4)

        for i in range(3,6):
            self.__create_list_with_choose_col_and_uniq_rows_and_rewrite_to_new_sheet(x=i,lev=3)

        for i in range(4,6):
            self.__create_list_with_choose_col_and_uniq_rows_and_rewrite_to_new_sheet(x=i,lev=2)



if __name__ == '__main__':

    f1 = Exl_data(500, 'данные_cust_segments.xlsx') #аргументы:N-клиентов и название файла.
    f1.level_6_all_cols()

    f1.savefile()  # тоже НЕ УБИРАТЬ! если хотите сохранения файла


