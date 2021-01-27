# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class Exl_data:

    def __init__(self,amount, name_file):  # обязательные  параметры при создании экземпляра
        self.amount = amount  # параметр для суммы клиентов
        self.name_file = name_file  # параметр для названия файла
        self.maxcol = 0  # макс.количство колонок в 1-ом листе
        self.maxrow = 0  # макс.количство строк в 1-ом листе
        self.count_row = 2  #
        self.find = 0  # сумма найденных строк
        self.clients = 0  # считаем сумму клиентов
        self.level = 0  # определим уровень комбинации поиска

        self.wb = load_workbook(filename=self.name_file)  # вставляем файл-эксель и присваеваем к wb
        self.sheet = self.wb.active.title  # текущий активный лист

        self.sh1 = self.wb['Сегменты']
        self.data2 = self.wb['Результат']['A1'].value  # 1-ый способ получить ячейку
        self.sh1.cell(6, 6).value  # 2-ой способ получить ячейку


        self.maxcol = self.sh1.max_column  # макс.количство колонок в 1-ом листе
        self.maxrow = self.sh1.max_row  # макс.количство строк в 1-ом листе
        print('amount- ',self.amount,', filename- ', self.name_file)
        print('maxcol- ',self.maxcol,', maxrow- ', self.maxrow)
        # print('sheets in file- ',self.wb.sheetnames)
        self.name_lev6 = ''
        self.pos_one = 1
        self.data_rows = []  # все строки 6-го уровня



    def __create_temp_sheet(self):
        self.wb.create_sheet(f'N={self.amount}')
        self.sh3 = self.wb[f'N={self.amount}']

    def __names_colons(self):

        temp_sh = self.wb[self.wb.sheetnames[-1]]
        for i in range(1, self.maxcol + 1):  # копируем самый первый ряд в новый лист
            col_titles = self.sh1.cell(1, i)  #
            temp_sh.cell(self.pos_one, i).value = col_titles.value  #
            temp_sh.cell(self.pos_one, i).fill = PatternFill("solid", fgColor="00FFCC99")  # красим

    def __find_amount_clients_in_all(self):
        for i in range(2, self.maxrow + 1):  # итерируем по количество строкам листа Сегменты со 2 строки,тк 1 это названия колонок.
            if self.sh1.cell(row=i, column=self.maxcol).value >= self.amount:  # тут будет N-аргумент из функции
                self.find += 1  # считаем найденные подходящие строки
                self.clients += self.sh1.cell(row=i, column=self.maxcol).value  # считаем всех клиентов у подход. строк
                for j in range(1, self.maxcol + 1):  # итерируем по количеству колонок в строке таблицы Сегменты
                    c = self.sh1.cell(row=i, column=j)  # подходящую строку присваиваем переменной с
                    self.sh3.cell(row=self.count_row, column=j).value = c.value  # копируем строку в другую таблицу
                self.count_row += 1  # считаем


    # тут будет скопировано с 6-го уровня строки
    def __create_list_from_lev6(self):  #  x - номер игнорируемой колонки (с 0 до 5)
        # count_clients_in_same_row = 0 # нулевая сумма клиентов каждой группы одинаковых строк
        # добавляем в список data_rows все строки из 6го уровня(лист sh3)
        for row in self.wb[f'N={self.amount} lev=6'].iter_rows(min_row=4, max_col=self.find+2, values_only=True):  # max_col это
            self.data_rows.append(list(row))


    # в data_same_rows уник.строки и подсчитаны суммы клиентов.
    # data_same_rows сохранится в конструкторе init
    # и будет доступен для других методов.
    # Метод позволяет выбрать колонку которую хотим закрыть
    def __create_list_with_choose_col_and_uniq_rows_and_rewrite_to_new_sheet(self,lev,x):
        data_same_rows = []

        # в списке data_row закрываем нужную нам колонку(ставим None).
        # Порядок элементов совпадает с порядком столбцов.
        # колонки с 0 до 5.  6-ая колонка это кол-во клиентов
        for i in self.data_rows:
            i[x] = 'XXX'

        # отсекаем одинаковые строки из data_row и
        # создаем список из уникальных строк в data_same_rows.
        for i in self.data_rows:
            if i[0:6] not in data_same_rows:
                data_same_rows.append(i[0:6])
        print(f'из level_{lev}_col_{x} найдено - ', len(data_same_rows))

        # в data_rows считаем сумму клиентов одинаковых строк и добавляем в уник.строку как
        # последний элемент в списке.
        # окончательное формирование списка уник.строк в data_same_rows
        for i in data_same_rows:
            count_clients = 0 # считаем клиентов всех одинаковых строк и вешаем на уникальную.
            for k in self.data_rows:
                if i == k[0:6]:
                    count_clients += k[6]
            i.append(count_clients)
        # print(self.data_same_rows)


    # переписываем уник.строки из data_same_rows в новый лист
    # def __rewrite_to_new_sheet(self,lev,x): #запись нового уровня в другой лист
        count_row_sh3 = 4 # точка отсчета rows в sh3(6-ой уровень)
        self.wb.create_sheet(f'N={self.amount} lev={lev} col={x}')  # создаем новый лист
        # temp_sheet = self.wb.sheetnames[-1]  # присваиваем новый лист к temp_sheet
        #self.wb[f'N={self.amount} lev=6']
        temp_sh = self.wb[f'N={self.amount} lev={lev} col={x}']
        len_data = len(data_same_rows)
        # name_main_level_sheet = self.wb[f'N={self.amount} lev=6']
        for spisok in range(0,len_data):

                for j in range(0, self.maxcol ):
                    cell = data_same_rows[spisok][j]
                    # cell = data_same_rows[count_elem][j]
                    # cell = self.sh3.cell(row=count_row_sh3, column=j)
                    # print(cell.value)
                    temp_sh.cell(row=count_row_sh3, column=j+1).value = cell
                count_row_sh3 += 1
        data_same_rows=[]
        self.data_rows = []

    def __show_end(self):
        reversed_find = range(self.find + 1, 0, -1)  # ставим обратный счет от числа найденных строк
        for row in reversed_find:  # с конца таблицы двигаем строки
            self.sh3.move_range(f"A{row}:G{row}", rows=2, cols=0)  # двигаем вниз 2 раза

        self.sh3.cell(self.pos_one, 1).value = f'N={self.amount}'  # сумма N из аргумента функции и красим
        self.sh3.cell(self.pos_one, 1).fill = PatternFill("solid", fgColor="00FFFF00")

        self.sh3.cell(self.pos_one, 2).value = f"S={self.clients}"  # сумма клиентов из аргумента функции и красим
        self.sh3.cell(self.pos_one, 2).fill = PatternFill("solid", fgColor="0000FF00")

        self.sh3.cell(self.pos_one, 3).value = 'level='  # укажем в ячейке уровень комбинации  и красим
        self.sh3.cell(self.pos_one, 3).fill = PatternFill("solid", fgColor="00FF99CC")



    def savefile(self):
        self.wb.save(f'amount_{self.amount}.xlsx')


    def level_6_all_cols(self):
        self.__create_temp_sheet()
        self.__names_colons()
        self.__find_amount_clients_in_all()
        print('Из level_6 найдено - ',self.find,'строк')
        self.__show_end()
        self.sh3.title = self.sh3.title + ' lev=6'


    def level_5_col_x(self,lev=5,x=1):  # lev = уровень в этом методе и колонка в этом методе и col по выбору

        # self.__create_temp_sheet()
        self.__create_list_from_lev6()
        self.__create_list_with_choose_col_and_uniq_rows_and_rewrite_to_new_sheet(lev,x)
        self.__names_colons()

    def level_4_col1_x_col2_y(self,lev=4,x=0,y=1):
        pass



if __name__ == '__main__':
    f1 = Exl_data(200, 'данные_cust_segments.xlsx')  # N-клиенты и название файла.
    f1.level_6_all_cols()  # НЕ УБИРАТЬ! тк все уровни делаются на основе этой.
    f1.level_5_col_x(x=0)  # 5-ый уровень, где х = номер колонки.
    f1.level_5_col_x(x=1)
    f1.level_5_col_x(x=2)
    f1.level_5_col_x(x=3)
    f1.level_5_col_x(x=4)
    f1.level_5_col_x(x=5)
    f1.savefile()  # сохранение файла



